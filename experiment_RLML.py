import logging
from IHML.utils import global_init as gi #Logging module initialized first
from IHML import globals as glb

from numpy import mean
from numpy import std
from sklearn.model_selection import train_test_split
from sklearn.metrics import f1_score, accuracy_score, recall_score,precision_score
import logging
import openpyxl  as openpyxl
import pandas as pd
import time
from train_RLML_agent import Train_IHML
from run_agent_RLML import Run_IHML
from IHML.incremental_metalearner import IncrementalMetalearner

#Global loger
import logging
logger = logging.getLogger()

class Experiment_IHML():

    def load_and_run_experiments(self):
        experiments = []
        wb = openpyxl.load_workbook(glb.EXPERIMENTS_INPUT_FILE)
        sheet = wb.active

        ## Create a dictionary of column names
        col_names = {}
        col_index = 0
        for col in sheet.iter_cols(1, sheet.max_column):
            col_names[col[0].value] = col_index
            col_index += 1

        # reading specific column
        params = []
        # params["ALPHA"]=sheet.cell(row=1, column=1).value
        # Read excel
        data = sheet.values
        # Get the first line in file as a header line
        columns = next(data)[0:]
        # Create a DataFrame based on the second and subsequent lines of data
        df = pd.DataFrame(data, columns=columns)
        for row_cells in sheet.iter_rows(min_row=2):
            if row_cells[col_names["STATE"]].value == "ready":
                experiment = {}
                ##Get list of parameters in dictionary format
                row_params = {}
                for col in col_names:
                    row_params[col] = row_cells[col_names[col]].value

                self.set_params(row_params)
                logger.info("EXPERIMENT NUMBER:"+str(glb.EXPERIMENT_ID))
                #Train the experiment dude
                experiement_start_time = time.time()
                train_IHML = Train_IHML()
                ihml = IncrementalMetalearner()
                XTrain, XTest, YTrain, YTest = ihml.load_dataset()
                train_IHML.train_ihml(XTrain, XTest, YTrain, YTest)

                #Run the experiment , test my boi
                logger.info("Starting RUN agent")
                run_agent_IHML = Run_IHML()
                outputs=run_agent_IHML.run_and_save_simulations(env=train_IHML.env, input_filename=glb.MODEL_DATA_FILE_NAME, n_episodes=glb.MAX_EPOCH)
                logger.info("RUN tests complete.")

                experiement_end_time = time.time()
                outputs["IHML_RL_Duration"] = experiement_end_time-experiement_start_time



                print("Updated experiment values:--")
                for col in columns:
                    if col in outputs:
                        row_cells[col_names[col]].value = outputs[col]

                # Mark the experiment complete
                row_cells[col_names["STATE"]].value = "done"

                wb.save(filename=glb.EXPERIMENTS_INPUT_FILE)

    def set_params(self,params):
        #Changeable params
        if "DATASET" in params: glb.DATASET = params["DATASET"]
        if "SCORER" in params:
            if params["SCORER"] == "accuracy":glb.SCORER=accuracy_score
            if params["SCORER"] == "f1": glb.SCORER = f1_score
            if params["SCORER"] == "recall":glb.SCORER=recall_score
            if params["SCORER"] == "precision": glb.SCORER = precision_score
        if "DATASET_SAMPLE_SIZE" in params: glb.DATASET_SAMPLE_SIZE = params["DATASET_SAMPLE_SIZE"]
        if "MAX_EPOCH" in params: glb.MAX_EPOCH = params["MAX_EPOCH"]
        if "EXPERIMENT_ID" in params: glb.EXPERIMENT_ID = params["EXPERIMENT_ID"]

        print("set experiment params complete")

    def run_experiment(self,X=None,Y=None):

        self.logger.info("New RL Experiment is running...")
        self.print_global_parameters()

        #Load data
        XTest, XTrain, YTest, YTrain = self.load_dataset()

        #Train the dude
        # Instantiate agent class
        logger.info("Started training...")



        self.fit(XTrain,YTrain,glb.DATASET)
        experiment_outputs= self.evaluate_experiment(XTest,YTest)
        ##Calculate AUC/ROC...
        ##ml.evaluate_model_roc(self,XTest,YTest)

        return experiment_outputs

        #Now we built model, final test
        #preds=self.predict(XTest)
        #print("> Test score: IncrementalMetaLearner: {} %".format(100.0 * glb.SCORER(YTest,preds)))

    def print_global_parameters(self):
        logger = logging.getLogger()
        logger.info("SCORER=" + str(glb.SCORER))
        logger.info("DATASET=" + str(glb.DATASET))
        logger.info("DO_SAMPLE=" + str(glb.DATASET_DO_SAMPLE))





experiments = Experiment_IHML()
experiments.load_and_run_experiments()


import warnings
warnings.filterwarnings("ignore")

logger.info("*****")
logger.info("RL EXPERIMENT RUNNER complete")
logger.info("*****")