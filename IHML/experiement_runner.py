#Logging module initialized first
import utils.global_init
from incremental_metalearner import IncrementalMetalearner
import logging
import globals as glb
import openpyxl  as openpyxl
import pandas as pd
import timeit

logger = logging.getLogger()
logger.info("*****")
logger.info("EXPERIMENT RUNNER started.")
logger.info("*****")

def load_experiments():
    experiments=[]
    wb = openpyxl.load_workbook(glb.EXPERIMENTS_FILE)
    sheet = wb.active


    ## Create a dictionary of column names
    col_names = {}
    col_index = 0
    for col in sheet.iter_cols(1, sheet.max_column):
        col_names[col[0].value] = col_index
        col_index += 1

    # reading specific column
    params=[]
    #params["ALPHA"]=sheet.cell(row=1, column=1).value
    # Read excel
    data = sheet.values
    # Get the first line in file as a header line
    columns = next(data)[0:]
    # Create a DataFrame based on the second and subsequent lines of data
    df = pd.DataFrame(data, columns=columns)
    for row_cells in sheet.iter_rows(min_row=2):
        print(str(row_cells))
        if row_cells[col_names["STATE"]].value=="ready":
            experiment = IncrementalMetalearner()
            ##Get list of parameters in dictionary format
            row_params={}
            for col in col_names:
                row_params[col] = row_cells[col_names[col]].value

            experiment.set_params(row_params)
            outputs = experiment.run_experiment()

            experiement_end_time = timeit.timeit()
            print("Updated experiment values:--")
            for col in columns:
                if col in outputs:
                        row_cells[col_names[col]].value = outputs[col]


            #Mark the experiment complete
            row_cells[col_names["STATE"]].value ="done"

            wb.save(filename=glb.EXPERIMENTS_FILE)


load_experiments()

# experiment = IncrementalMetalearner()
# experiment.set_params({"ALPHA":20,"SCORE_IMPROVEMENT_TYPE":"Any"})
# experiment.run_experiment()


import warnings
warnings.filterwarnings("ignore")

logger.info("*****")
logger.info("EXPERIMENT RUNNER complete")
logger.info("*****")






